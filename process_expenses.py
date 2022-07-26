from datetime import datetime
import xlsxwriter
import openpyxl
import shutil
import os


class ProcessExpenses:
    def __init__(self, new_expense=None, reset=None, test=True):
        if test:
            self.base_folder = os.path.dirname(os.path.abspath(__file__))
        else:
            self.base_folder = 'C:\\Users\Tom\Important\Self Employed\Yodel'
        self.file_name = 'CutlerT Expenses 2021-22.xlsx'
        self.file_path = os.path.join(self.base_folder, self.file_name)
        self.copy_folder = os.path.join(self.base_folder, 'Expenses 2021-22 Copies')
        self.file_header = [['Date', 12], ['Amount', 10], ['Reason', 12], ['Note', 12]]
        self.new_expense = new_expense
        self.reset = reset
        self.date_order = []
        self.expense_data = {}
        self.process_expenses()

    def process_expenses(self):
        if self.reset:
            self.copy_file_to(os.path.join(self.copy_folder, f'({self.reset}) ' + self.file_name), self.file_path)
        else:
            self.copy_file_to(self.file_path, os.path.join(self.copy_folder, f'({len(os.listdir(self.copy_folder)) + 1}) ' + self.file_name))
            if self.new_expense:
                self.get_expense_data()
                self.add_expense_date()
                self.write_output_file()

    def copy_file_to(self, file, to):
        if os.path.exists(file):
            shutil.copy(file, to)

    def get_expense_data(self):
        counter = 0
        for row in openpyxl.load_workbook(self.file_path).active.iter_rows(min_row=2):
            date = row[0].value
            if date:
                self.expense_data[counter] = [date]
                for cell in row[1:len(self.file_header)]:
                    self.expense_data[counter].append(cell.value)
                counter += 1
            else:
                break
        self.date_order = [*range(counter)]

    def add_expense_date(self):
        for new_expense in self.new_expense:
            new_key = len(self.expense_data.keys())
            self.expense_data[new_key] = [datetime.strptime(new_expense[0], '%d/%m/%y').strftime('%d-%b-%Y'), new_expense[1], new_expense[2], new_expense[3] if len(new_expense) == 4 else None]
            self.date_order.append(new_key)
            for index, key in enumerate(self.date_order):
                if datetime.strptime(self.expense_data[key][0], '%d-%b-%Y') > datetime.strptime(new_expense[0], '%d/%m/%y'):
                    self.date_order.remove(new_key)
                    self.date_order.insert(index, new_key)
                    break

    def write_output_file(self):
        workbook = xlsxwriter.Workbook(self.file_path)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold': True})
        right = workbook.add_format({'align': 'right'})
        currency = workbook.add_format({'num_format': '£#,##0.00'})
        for col, header in enumerate(self.file_header):
            worksheet.write(0, col, header[0], bold)
            worksheet.set_column(col, col + 1, header[1])
        for row, data in enumerate(self.date_order):
            for col, cell in enumerate(self.expense_data[data]):
                worksheet.write(row + 1, col, cell, bold if col == 0 else (currency if not isinstance(cell, str) else right))
        workbook.close()


if __name__ == '__main__':
    ProcessExpenses()
    # ProcessExpenses(new_expense=[['25/07/22', 90.38, 'Fuel'], ['18/07/22', 97.80, 'Fuel'], ['02/07/22', 89.96, 'Fuel'], ['27/06/22', 79.88, 'Fuel'], ['20/06/22', 98.27, 'Fuel'], ['14/06/22', 79.74, 'Fuel'],
    #                              ['13/06/22', 679.51, 'Service'], ['24/06/22', 21.07, 'Injector Cleaner'], ['19/06/22', 12.46, 'Pens'], ['04/07/22', 25.69, 'Phone Bill'], ['25/07/22', 25.69, 'Phone Bill']])
    # ProcessExpenses(new_expense=[[date (dd/mm/yy) {str}, amount {int}, reason {str}, note (optional) {str}]])
    # ProcessExpenses(reset=1)

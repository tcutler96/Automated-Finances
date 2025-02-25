import os
import json
import shutil
from math import ceil
from datetime import datetime
from openpyxl import load_workbook, styles


def get_base_path(test=True):
    if test:
        return os.path.abspath('..')
    else:
        return 'C:\\Users\Tom\Important\Self Employed\Yodel'


def copy_file(file, dest):
    if os.path.isfile(file) & os.path.exists(file):
        shutil.copy(file, dest)


def get_email_credentials():
    with open('email_credentials.json', 'r') as f:
        file = json.load(f)
        email_username = file['Email Username']
        email_password = file['Email Password']
    return email_username, email_password


def get_financial_year(date):
    date = datetime.strptime(date, '%d-%b-%Y')
    for year in range(max(datetime.today().year, date.year), 2020, -1):
        financial_year = [datetime(year=year - 1, month=4, day=6), datetime(year=year, month=4, day=5)]
        if financial_year[0] <= date <= financial_year[1]:
            return f'{year - 1}-{year}', f'06-Apr-{year - 1}', f'05-Apr-{year}'


def get_fortnights(date_from, date_to):
    date_from = datetime.strptime(date_from, '%d-%b-%Y')
    date_to = datetime.strptime(date_to, '%d-%b-%Y')
    return ceil((date_to - date_from).days / 14)


class Workbook:
    def __init__(self, path):
        self.path = path
        self.workbook = load_workbook(path)
        self.worksheet = None
        self.formats = {'bold': styles.Font(bold=True),
                        'right': styles.Alignment(horizontal='right'),
                        'money': '£##,#0.00'}

    def set_worksheet(self, name):
        if name and name in self.workbook.sheetnames:
            self.worksheet = self.workbook[name]
            self.workbook.active = self.workbook[name]

    def add_worksheet(self, name, header_data=None):
        if name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=name)
            self.set_worksheet(name=name)
            if header_data:
                for col, header in enumerate(header_data):
                    cell = self.worksheet.cell(row=1, column=col + 1)
                    cell.value = header[0]
                    cell.font = styles.Font(bold=True)
                    self.worksheet.column_dimensions[cell.column_letter].width = header[1]

    def write_row(self, row, values, insert=False):
        if insert:
            self.worksheet.insert_rows(row)
        for col, value in enumerate(values):
            if value[0]:
                self.write_cell(row=row, col=col + 1, value=value[0], form=value[1])

    def write_cell(self, row, col, value, form=None):
        self.worksheet.cell(row=row, column=col).value = value
        if form:
            self.format_cell(self.worksheet.cell(row=row, column=col), form)

    def format_cell(self, cell, form=None):
        if form == 'bold':
            cell.font = self.formats['bold']
        elif form == 'right':
            cell.alignment = self.formats['right']
        elif form == 'money':
            cell.number_format = self.formats['money']

    def save_workbook(self, path):
        self.workbook.active = self.workbook[self.workbook.sheetnames[0]]
        self.workbook.save(path)
